import logging
import os
import re
from aiogram import Bot, Dispatcher, types
from aiogram.contrib.middlewares.logging import LoggingMiddleware
from flask import Flask, render_template, send_from_directory, request, jsonify
import threading
from jinja2 import Environment

import sqlite3
import asyncio

from openpyxl.workbook import Workbook
from difflib import get_close_matches

from TOKEN_name import token
import pandas as pd
from openpyxl import load_workbook
from aiogram.dispatcher import FSMContext
from aiogram.dispatcher.filters.state import State, StatesGroup
from aiogram.contrib.fsm_storage.memory import MemoryStorage

# Устанавливаем уровень логгирования
logging.basicConfig(level=logging.INFO)

TOKEN = token
DATABASE_FILE = 'database.db'

# Инициализируем бота и диспетчера
storage = MemoryStorage()
bot = Bot(token=TOKEN, timeout=30)
dp = Dispatcher(bot, storage=storage)
dp.middleware.setup(LoggingMiddleware())
db = sqlite3.connect(DATABASE_FILE, check_same_thread=False)

cache = {}  # Простой словарь для кэширования chat_id

# Инициализируем Flask приложение
app = Flask(__name__)

# Создаем экземпляр окружения Jinja2
env = Environment()

# Флаг для отслеживания запущен ли бот
bot_running = False
# Чтение данных из файла Excel
df = pd.read_excel(r'C:\Users\FR\Desktop\BD_film.xlsx')
# Преобразование данных в список словарей
films = df.to_dict(orient='records')

EXCEL_FILE_PATH = "C:\\Users\\FR\\Desktop\\BD_film.xlsx"


class MyStates(StatesGroup):
    waiting_for_message = State()
    waiting_for_action = State()


# Класс для определения состояний FSM
class AddFilmStates(StatesGroup):
    WAITING_FOR_VIDEO = State()
    WAITING_FOR_NAME = State()
    WAITING_FOR_DESCRIPTION = State()
    WAITING_FOR_RELEASE_DATE = State()
    WAITING_FOR_RATING = State()
    WAITING_FOR_POSTER = State()


# Функция для сохранения данных в Excel
async def save_to_excel(column, value):
    try:
        workbook = load_workbook(filename=EXCEL_FILE_PATH)
    except FileNotFoundError:
        workbook = Workbook()
    sheet = workbook.active
    # Находим первую свободную строку в указанном столбце
    row_number = 1
    while sheet.cell(row=row_number, column=column).value is not None:
        row_number += 1
    # Сохраняем значение в указанном столбце
    sheet.cell(row=row_number, column=column, value=value)
    workbook.save(EXCEL_FILE_PATH)


async def create_table():
    with db:
        cursor = db.cursor()
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS users (
                user_id INTEGER PRIMARY KEY,
                active INTEGER DEFAULT 1
            )
        ''')
        db.commit()


async def user_exists(user_id):
    with db:
        cursor = db.cursor()
        result = cursor.execute("SELECT * FROM users WHERE user_id = ?", (user_id,)).fetchone()
        return bool(result)


async def add_user(user_id):
    with db:
        cursor = db.cursor()
        try:
            print(f"Добавление пользователя с user_id {user_id}")
            cursor.execute("INSERT OR IGNORE INTO users (user_id) VALUES (?)", (user_id,))
            db.commit()

            last_row_id = cursor.lastrowid

            if last_row_id > 0:
                print(f"Пользователь {user_id} успешно добавлен с ID {last_row_id}.")
                await set_active(user_id, 1)
                return user_id
            else:
                print(f"Пользователь с user_id {user_id} уже существует.")
                await set_active(user_id, 1)
                return user_id

        except Exception as e:
            print(f"Произошла ошибка при добавлении/обновлении пользователя. {e}")
            return None


async def set_active(user_id, active):
    with db:
        cursor = db.cursor()
        try:
            current_status = cursor.execute("SELECT active FROM users WHERE user_id = ?", (user_id,)).fetchone()

            if current_status and current_status[0] != active:
                all_users_before = cursor.execute("SELECT * FROM users").fetchall()
                print("Все пользователи в базе данных перед обновлением:", all_users_before)

                cursor.execute("UPDATE users SET active = ? WHERE user_id = ?", (active, user_id))
                db.commit()

                all_users_after = cursor.execute("SELECT * FROM users").fetchall()
                print("Все пользователи в базе данных после обновления:", all_users_after)

                print(f"Статус пользователя {user_id} успешно обновлен на {active}")
                return 1  # Возвращаем 1, так как одна запись была изменена
            else:
                print(f"Статус пользователя {user_id} уже установлен в {active}")
                return 0
        except Exception as e:
            print(f"Произошла ошибка при обновлении статуса пользователя {user_id}. {e}")
            return 0


async def get_users():
    with db:
        cursor = db.cursor()
        return cursor.execute("SELECT user_id, active FROM users").fetchall()


async def get_all_users():
    with db:
        cursor = db.cursor()
        cursor.execute("SELECT user_id, active FROM users ORDER BY user_id")
        users = []
        row = cursor.fetchone()
        while row is not None:
            users.append(row)
            row = cursor.fetchone()
        return users


# Определяем функцию для загрузки фильмов из Excel файла
def load_films_from_excel(excel_file_path):
    try:
        df = pd.read_excel(excel_file_path)
        films = df.to_dict(orient='records')
        return films
    except FileNotFoundError:
        # Если файл не найден, вернуть пустой список
        return []
    except Exception as e:
        # Обработка других исключений при чтении Excel-файла
        print(f"Error loading films from Excel: {e}")
        return []


# Определяем фильтр shuffle_films
def shuffle_films(films):
    import random
    shuffled_films = list(films)
    random.shuffle(shuffled_films)
    return shuffled_films


# Регистрируем фильтр в окружении Jinja2
env.filters['shuffle_films'] = shuffle_films


def clean_text(text):
    """Очищает текст от спецсимволов и приводит к нижнему регистру."""
    if isinstance(text, str):
        return re.sub(r'\W+', '', text.lower())
    else:
        return str(text)


def load_file_ids_from_excel():
    desktop_path = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')
    excel_file_path = os.path.join(desktop_path, 'BD_film.xlsx')

    try:
        workbook = load_workbook(filename=excel_file_path)
    except FileNotFoundError:
        workbook = Workbook()
        workbook.save(excel_file_path)

    sheet = workbook.active
    file_ids = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if len(row) >= 2:
            track_name, file_id = row[:2]
            file_ids[track_name] = file_id
    return file_ids


FILE_ID_STORAGE = load_file_ids_from_excel()
SENT_TRACKS = {}

# Загружаем данные из Excel при старте бота
VIDEO_FILE_IDS = {}


def load_video_file_ids_from_excel():
    try:
        workbook = load_workbook(filename="C:\\Users\\FR\\Desktop\\BD_film.xlsx")
    except FileNotFoundError:
        workbook = Workbook()
        workbook.save("C:\\Users\\FR\\Desktop\\BD_film.xlsx")

    sheet = workbook.active
    video_file_ids = {}
    for row in sheet.iter_rows(values_only=True):
        if len(row) >= 3:
            video_name, _, file_id, *_ = row
            if video_name is not None:
                video_name_cleaned = clean_text(video_name)

                if video_name_cleaned in video_file_ids:
                    video_file_ids[video_name_cleaned].append(file_id)
                else:
                    video_file_ids[video_name_cleaned] = [file_id]
    return video_file_ids


def web_app_keyboard():
    keyboard = types.ReplyKeyboardMarkup(row_width=1)
    web_app_test = types.WebAppInfo("https://telegram.mihailgok.ru")
    web_app_game = types.WebAppInfo("https://games.mihailgok.ru")
    one_butt = types.KeyboardButton(text="Тестовая страница", web_app=web_app_test)
    two_butt = types.KeyboardButton(text="Игра", web_app=web_app_game)
    keyboard.add(one_butt, two_butt)

    return keyboard


@dp.зmessage_handler(commands=['start'])
async def start_command(message: types.Message):
    if message.chat.type == 'private':
        user_id = message.from_user.id
        await create_table()
        if not await user_exists(user_id):
            await add_user(user_id)
        else:
            await set_active(user_id, 1)
        await bot.send_message(user_id, 'Добро пожаловать')

    # Отправляем приветственное сообщение и запоминаем chat_id
    chat_id = message.chat.id
    await bot.send_message(chat_id=chat_id, text='Привет! Я простой бот.')
    # Получаем chat_id пользователя
    chat_id = message.chat.id
    # Сохраняем chat_id в кэше
    cache[chat_id] = True
    await message.reply(f"Ваш chat_id: {chat_id}")


@dp.message_handler(content_types="web_app_data")
async def answer(web_app_mes: types.Message):
    print(web_app_mes)
    print(web_app_mes.web_app_data.data)
    await web_app_mes.answer(f"получили информацию из веб-приложения: {web_app_mes.web_app_data.data}")


@dp.message_handler(commands=["message_to_all_users_from_admin"])
async def command_message_to_all_users_from_admin(message: types.Message):
    user_id = message.from_user.id

    # Проверяем, является ли пользователь rene_sjelen
    if user_id != 374208152:  # Замените 123456789 на фактический user_id пользователя rene_sjelen
        await bot.send_message(user_id, "У вас нет прав для выполнения этой команды.")
        return

    await MyStates.waiting_for_message.set()
    await bot.send_message(user_id, "Введите сообщение для рассылки:")


@dp.message_handler(state=MyStates.waiting_for_message)
async def process_message_to_all_users_from_admin(message: types.Message, state: FSMContext):
    user_id = message.from_user.id
    await state.update_data(message_text=message.text)
    await MyStates.waiting_for_action.set()
    await bot.send_message(user_id, f"Вы ввели следующее сообщение:\n\n{message.text}\n\n"
                                    "Выберите вариант действия:\n"
                                    "1. Отправить это всем\n"
                                    "2. Отменить отправку")


@dp.message_handler(state=MyStates.waiting_for_action)
async def process_action_for_all_users(message: types.Message, state: FSMContext):
    user_id = message.from_user.id
    data = await state.get_data()
    message_text = data.get('message_text', '')
    action = message.text.strip().lower()

    try:
        if action == '1':
            all_users = await get_all_users()
            for user in all_users:
                user_id, active = user
                if active == 1:
                    await bot.send_message(user_id, f"📣 Важное объявление от администратора:\n\n{message_text}")

            # Проверяем, является ли пользователь с идентификатором 374208152
            if user_id == 374208152:
                await bot.send_message(374208152, "Сообщение успешно отправлено всем активным пользователям.")
        elif action == '2':
            await bot.send_message(user_id, "Отправка отменена.")
        else:
            await bot.send_message(user_id, "Неверный выбор. Пожалуйста, введите 1 или 2.")
    except Exception:
        # Обработка блокировки пользователя
        print(f"Пользователь {user_id} заблокировал бота. Обновляем статус в базе данных.")
        await set_active(user_id, 0)
    finally:
        await state.finish()


# Обработчик добавления нового фильма
@dp.message_handler(commands=['add_new_film'])
async def add_film_command(message: types.Message):
    # Проверяем, что отправитель - пользователь @rene_sjelen
    if message.from_user.username == 'rene_sjelen':
        await message.answer("Отправьте видеофайл для добавления:")
        # Переход в состояние ожидания ввода видеофайла
        await AddFilmStates.WAITING_FOR_VIDEO.set()
    else:
        await message.answer("Извините, у вас нет разрешения на выполнение этой команды.")


# Обработчик добавления видеофайла
@dp.message_handler(state=AddFilmStates.WAITING_FOR_VIDEO, content_types=types.ContentTypes.VIDEO)
async def add_film_video(message: types.Message, state: FSMContext):
    # Убедимся, что в сообщении есть видео
    if message.video is not None:
        # Получим информацию о видео
        file_id = message.video.file_id
        file_name = message.video.file_name
        # Сохраним идентификатор файла и оригинальное название файла в документе Excel
        await save_to_excel(2, file_name)  # Сохраняем оригинальное название файла в столбец B
        await save_to_excel(3, file_id)  # Сохраняем идентификатор файла в столбец C
        # Переход в состояние ожидания ввода названия фильма
        await AddFilmStates.WAITING_FOR_NAME.set()
        # Попросим пользователя ввести название фильма
        await message.answer("Теперь введите название фильма:")
    else:
        # Если видео в сообщении не найдено, уведомим пользователя
        await message.answer("В запросе не найдено видео. Пожалуйста, отправьте видеофайл.")


# Обработчик добавления названия фильма
@dp.message_handler(state=AddFilmStates.WAITING_FOR_NAME)
async def add_film_name(message: types.Message, state: FSMContext):
    film_name = message.text
    await save_to_excel(1, film_name)  # Сохраняем название фильма в столбец B
    await state.update_data(film_name=film_name)
    print("Film Name saved to FSMContext:", film_name)  # Добавляем отладочное сообщение
    await message.answer("Теперь введите год выхода фильма:")
    await AddFilmStates.WAITING_FOR_RELEASE_DATE.set()


# Обработчик добавления даты выхода фильма
@dp.message_handler(state=AddFilmStates.WAITING_FOR_RELEASE_DATE)
async def add_film_release_date(message: types.Message, state: FSMContext):
    release_date = message.text
    await save_to_excel(5, release_date)  # Сохраняем дату выпуска фильма в столбец E
    await state.update_data(release_date=release_date)
    print("Release Date saved to FSMContext:", release_date)  # Добавляем отладочное сообщение
    await message.answer("Теперь введите рейтинг фильма:")
    await AddFilmStates.WAITING_FOR_RATING.set()  # Переход к ожиданию ввода рейтинга


# Обработчик добавления рейтинга фильма
@dp.message_handler(state=AddFilmStates.WAITING_FOR_RATING)
async def add_film_rating(message: types.Message, state: FSMContext):
    rating = message.text
    await save_to_excel(7, rating)  # Сохраняем рейтинг фильма в столбец G
    await state.update_data(rating=rating)  # Сохраняем рейтинг в состоянии
    print("Rating saved to FSMContext:", rating)  # Добавляем отладочное сообщение
    await message.answer("Теперь введите описание фильма:")
    await AddFilmStates.WAITING_FOR_DESCRIPTION.set()  # Переход к ожиданию ввода описания фильма


# Обработчик добавления описания фильма
@dp.message_handler(state=AddFilmStates.WAITING_FOR_DESCRIPTION)
async def add_film_description(message: types.Message, state: FSMContext):
    description = message.text
    await save_to_excel(4, description)  # Сохраняем описание фильма в столбец D
    await state.update_data(description=description)
    print("Description saved to FSMContext:", description)  # Добавляем отладочное сообщение
    # Запрашиваем у пользователя отправку изображения
    await AddFilmStates.WAITING_FOR_POSTER.set()
    await message.answer("Теперь отправьте картинку фильма:")


# Обработчик добавления постера фильма
@dp.message_handler(state=AddFilmStates.WAITING_FOR_POSTER, content_types=types.ContentTypes.PHOTO)
async def add_film_poster(message: types.Message, state: FSMContext):
    photo = message.photo[-1]  # Берем последнее фото (самое качественное)
    file_name = f"{message.from_user.id}_{photo.file_unique_id}.jpg"
    save_path = f"C:\\Users\\FR\\PycharmProjects\\pythonProject\\static\\posters\\{file_name}"
    await photo.download(destination_file=save_path)  # Скачиваем фото
    poster_path = f"posters/{file_name}"  # Формируем относительный путь для сохранения в Excel
    await save_to_excel(6, poster_path)  # Сохраняем путь к постеру в столбец F

    # Извлекаем данные из состояния
    data = await state.get_data()
    film_name = data.get('film_name')
    release_date = data.get('release_date')
    description = data.get('description')
    rating = data.get('rating')

    # Проверяем, что данные успешно извлечены
    print("Film Name:", film_name)
    print("Release Date:", release_date)
    print("Description:", description)
    print("rating:", rating)

    # Формируем сообщение
    response_message = (
        f'<b>«{film_name}» ({release_date})</b> Рейтинг: <b>{rating}</b>\n\n'
        f'{description}\n\n'
        f'Доступен в боте: @ParcerClub34_bot\n'
        f'По запросу: <code><strong>{film_name}</strong></code>'
    )
    with open(save_path, 'rb') as photo_file:
        await message.answer_photo(photo=photo_file, caption=response_message, parse_mode='HTML')

    await state.finish()  # Завершаем состояние FSMContext


@dp.message_handler(commands=['show_global_variables'])
async def show_global_variables(message: types.Message):
    await message.answer(f"{global_chat_id}")
    await message.answer(f"{global_fail_id}")


@dp.message_handler(commands=['send_video_global_user_chat'])
async def start_send_video_id(message: types.Message = None):  # Передаем аргумент по умолчанию
    chat_id = global_chat_id
    fail_id = global_fail_id
    if message:  # Проверяем, что message не None, чтобы избежать ошибки при ответе
        await message.answer("Спасибо, данные получены, ловите видео!")
    await bot.send_video(chat_id, fail_id)


@dp.message_handler(lambda message: message.from_user.username == "rene_sjelen", commands=['reload'])
async def reload_data(message: types.Message):
    global VIDEO_FILE_IDS
    try:
        VIDEO_FILE_IDS = load_video_file_ids_from_excel()
        await message.answer("Данные успешно обновлены!")
        print("Данные успешно обновлены!")
    except Exception as e:
        print(f"Ошибка при обновлении данных: {e}")
        await message.answer(f"Ошибка при обновлении данных: {e}")
        # Попробуйте добавить следующие строки для более подробного вывода ошибок
        import traceback
        traceback.print_exc()


@dp.message_handler()
async def handle_user_input(message: types.Message):
    if message.chat.type == 'private':
        user_id = message.from_user.id
        await set_active(user_id, 1)  # Обновляем статус пользователя на 1 для любого сообщения
        print(f"Статус пользователя {user_id} обновлен на 1")

        # Получаем актуальные данные из базы данных после обновления статуса
        all_users = await get_all_users()
        print("Все пользователи и их статус после обновления:")
        for user in all_users:
            user_id, active = user
            print(f"Пользователь {user_id}: Активен - {active}")

    video_name = clean_text(message.text)
    matching_names = get_close_matches(video_name, VIDEO_FILE_IDS.keys(), n=5, cutoff=0.6)

    if matching_names:
        for name in matching_names:
            file_ids = VIDEO_FILE_IDS[name]
            for file_id in file_ids:
                description, film_name = await get_description_and_name_from_excel(name)
                year = await get_year_from_excel(name)
                if description and film_name:
                    await send_video_to_user(message.chat.id, file_id, film_name, description, year)


async def send_video_to_user(chat_id, file_id, video_name, description, year):
    bold_video_name = f'<b>{video_name}</b>'
    video_caption = f'{bold_video_name} ({year}) @ParcerClub34_bot\n\n{description}\n\nБольше фильмов: @biblio_anime'
    await bot.send_video(chat_id, file_id, caption=video_caption, parse_mode='HTML')


async def get_description_and_name_from_excel(video_name):
    try:
        workbook = load_workbook(filename="C:\\Users\\FR\\Desktop\\BD_film.xlsx")
    except FileNotFoundError:
        return None, None

    sheet = workbook.active
    for row in sheet.iter_rows(min_row=2, values_only=True):  # Начинаем со второй строки
        if len(row) >= 7:  # Убедитесь, что в строке достаточно элементов
            name, _, _, description, _, _, _ = row  # Распаковка кортежа с учетом всех столбцов
            if name and clean_text(name) == video_name:
                return description, name

    return None, None


async def get_year_from_excel(video_name):
    try:
        workbook = load_workbook(filename="C:\\Users\\FR\\Desktop\\BD_film.xlsx")
    except FileNotFoundError:
        return None

    sheet = workbook.active
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if len(row) >= 7:  # Проверка на количество элементов в строке
            name, _, _, _, year, _, _ = row  # Игнорируем лишние значения
            if name and clean_text(name) == video_name:
                return year

    return None


df = pd.read_excel(r'C:\Users\FR\Desktop\BD_film.xlsx')
films_data = df.to_dict(orient='records')

# Путь к SSL-сертификату и ключу
cert_path = r'C:\Certbot\live\afishakinobota.site\fullchain.pem'
key_path = r'C:\Certbot\live\afishakinobota.site\privkey.pem'

context = (cert_path, key_path)


# Функция для запуска Flask в отдельном потоке

def run_flask_app():
    app.run(host='0.0.0.0', ssl_context=context, port='443', debug=True, use_reloader=False, threaded=True)
    #


# Использование функции в коде Flask
@app.route('/')
def start_command_route():
    # Используем render_template для отображения страницы из файла index.html
    return render_template('index.html', films=load_films_from_excel(EXCEL_FILE_PATH), shuffle_films=shuffle_films)


@app.route('/film/<fail_id>')
def film_page(fail_id):
    film_data = get_film_data_by_id(fail_id)

    if film_data:
        return render_template('film_page.html', film_data=film_data)
    else:
        return render_template('not_found.html')


# Определение маршрута для обработки запросов по адресу /.well-known/pki-validation/<filename>
@app.route('/.well-known/acme-challenge/<path:filename>')
def serve_text_file(filename):
    # Абсолютный путь к папке, где находится текстовый файл
    directory = 'C:/Users/FR/PycharmProjects/pythonProject/key'

    # Возвращаем содержимое текстового файла
    return send_from_directory(directory, filename)


# Объявляем глобальные переменные
global_chat_id = None
global_fail_id = None


async def send_video(chat_id, fail_id):
    try:
        await bot.send_video(chat_id, fail_id)
        return "Video sent successfully"
    except Exception as e:
        return f"Error sending video: {e}", 500


@app.route('/watch-film', methods=['POST'])
async def watch_film():
    # Получаем данные из запроса
    data = await request.get_json()
    chat_id = data.get('userId')
    fail_id = data.get('failId')

    # Логируем информацию
    app.logger.info(f"Пользователь с ID {chat_id}. Fail ID фильма: {fail_id}")

    # Вызываем функцию отправки видео из асинхронного контекста
    result = await send_video(chat_id, fail_id)
    return result


def get_film_data_by_id(fail_id):
    for film in films_data:
        if film.get('Fail-ID_film') == fail_id:
            return film

    return None


# Функция для on_startup
async def on_startup(dp):
    global bot_running
    print('Бот запущен!')
    bot_running = True


# Запуск бота и Flask с использованием executor
if __name__ == '__main__':
    VIDEO_FILE_IDS = load_video_file_ids_from_excel()
    from aiogram import executor as aiogram_executor

    # Запуск Flask в отдельном потоке (если бот еще не запущен)
    if not bot_running:
        threading.Thread(target=run_flask_app).start()

    # Запуск aiogram executor
    aiogram_executor.start_polling(dp, on_startup=on_startup, skip_updates=True)

    loop = asyncio.get_event_loop()
    loop.create_task(app.run_async(debug=True))
    loop.run_forever()
